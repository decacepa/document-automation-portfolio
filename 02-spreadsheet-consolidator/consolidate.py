"""
Spreadsheet consolidator.

Merges every individual claim into one master workbook and, more importantly,
refuses to hide a bad row. Each row is checked against the rules in
config.json; anything that fails lands on an Issues sheet with the file, the
cell and the reason, so a human fixes the source instead of the total.

Totals in the master are live formulas, not values computed in Python — the
sheet keeps working after someone edits it.

Usage:
    python consolidate.py
"""

import json
import logging
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill

BASE = Path(__file__).parent
HEADERS = ["Source file", "Staff ID", "Name", "Date", "Route", "Trips", "Unit cost", "Total"]


def load_config(path=BASE / "config.json"):
    with open(path, encoding="utf-8") as f:
        return json.load(f)


def setup_logging(logfile):
    logfile = Path(logfile)
    logfile.parent.mkdir(parents=True, exist_ok=True)
    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s  %(levelname)-7s %(message)s",
        handlers=[logging.FileHandler(logfile, encoding="utf-8"), logging.StreamHandler()],
    )


def read_claim(path, cfg, issues):
    """Return the valid rows of one claim file, appending any problems to issues."""
    wb = load_workbook(path, data_only=True)
    ws = wb.active

    staff_id = ws[cfg["source"]["staff_id_cell"]].value
    name = ws[cfg["source"]["name_cell"]].value
    first_row = cfg["source"]["first_data_row"]

    if not staff_id:
        issues.append([path.name, cfg["source"]["staff_id_cell"], "missing staff ID — file skipped"])
        return [], None

    rows = []
    seen = set()

    for r in range(first_row, ws.max_row + 1):
        date, route, trips, cost = (ws.cell(row=r, column=c).value for c in range(1, 5))

        if date is None and route is None:
            continue  # end of the table or a spacer row
        if isinstance(date, str) and date.strip().lower().startswith("claim total"):
            break

        if trips is None or cost is None:
            issues.append([path.name, f"row {r}", "blank value in Trips or Unit cost"])
            continue

        if not isinstance(trips, (int, float)) or not isinstance(cost, (int, float)):
            issues.append([path.name, f"row {r}", "number stored as text — will not sum"])
            continue

        key = (str(date), route, trips, cost)
        if key in seen:
            issues.append([path.name, f"row {r}", "duplicate of an earlier row"])
            continue
        seen.add(key)

        if cost > cfg["validation"]["max_unit_cost"]:
            issues.append([path.name, f"row {r}", f"unit cost above limit ({cost})"])
            continue

        rows.append([path.name, staff_id, name, str(date), route, trips, cost])

    return rows, staff_id


def write_master(rows, issues, out_path, cfg):
    wb = Workbook()
    ws = wb.active
    ws.title = "Consolidated"

    header_font = Font(name="Arial", size=10, bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="2E3849")

    for col, header in enumerate(HEADERS, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    for i, row in enumerate(rows, start=2):
        for col, value in enumerate(row, start=1):
            ws.cell(row=i, column=col, value=value).font = Font(name="Arial", size=10)
        # Live formula, so the master recalculates if a value is edited later.
        ws.cell(row=i, column=8, value=f"=F{i}*G{i}").font = Font(name="Arial", size=10)

    last = len(rows) + 1
    total_row = last + 2
    ws.cell(row=total_row, column=7, value="TOTAL").font = Font(name="Arial", size=10, bold=True)
    ws.cell(row=total_row, column=8, value=f"=SUM(H2:H{last})").font = Font(name="Arial", size=10, bold=True)
    ws.cell(row=total_row + 1, column=7, value="Rows").font = Font(name="Arial", size=10)
    ws.cell(row=total_row + 1, column=8, value=f"=COUNT(H2:H{last})").font = Font(name="Arial", size=10)

    for column, width in zip("ABCDEFGH", (26, 14, 20, 12, 12, 8, 11, 12)):
        ws.column_dimensions[column].width = width
    ws.freeze_panes = "A2"

    iss = wb.create_sheet("Issues")
    for col, header in enumerate(["Source file", "Location", "Problem"], start=1):
        cell = iss.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill

    for i, issue in enumerate(issues, start=2):
        for col, value in enumerate(issue, start=1):
            iss.cell(row=i, column=col, value=value).font = Font(name="Arial", size=10)

    if not issues:
        iss.cell(row=2, column=1, value="No issues found.").font = Font(name="Arial", size=10)

    for column, width in zip("ABC", (26, 14, 46)):
        iss.column_dimensions[column].width = width

    wb.save(out_path)


def main():
    cfg = load_config()
    setup_logging(BASE / cfg["paths"]["log_file"])

    folder = BASE / cfg["paths"]["individual"]
    out_path = BASE / cfg["paths"]["master_file"]
    files = sorted(folder.glob("*.xlsx"))

    if not files:
        logging.warning("no files in %s — run generate_samples.py first", folder)
        return

    started = datetime.now()
    logging.info("consolidating %d files", len(files))

    all_rows, issues, staff = [], [], set()
    for path in files:
        rows, staff_id = read_claim(path, cfg, issues)
        all_rows.extend(rows)
        if staff_id:
            staff.add(staff_id)
        logging.info("%-28s %3d valid rows", path.name, len(rows))

    write_master(all_rows, issues, out_path, cfg)

    elapsed = (datetime.now() - started).total_seconds()
    logging.info("master written to %s", out_path.name)
    logging.info("%d rows from %d people · %d issues flagged · %.1fs",
                 len(all_rows), len(staff), len(issues), elapsed)
    if issues:
        logging.warning("issues need a human — see the Issues sheet")


if __name__ == "__main__":
    main()
