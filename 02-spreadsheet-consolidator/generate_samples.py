"""
Generate the individual monthly sheets used to demonstrate the consolidator.

Fictional data only. Three files are seeded with the defects that show up in
real submissions: a missing value, a number typed as text, and a duplicated
entry. The consolidator has to catch all three.

Usage:
    python generate_samples.py
"""

import random
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font

INDIVIDUAL = Path(__file__).parent / "individual"

FIRST = ["Alex", "Jordan", "Morgan", "Casey", "Riley", "Taylor", "Jamie", "Avery", "Quinn", "Reese",
         "Devon", "Skyler"]
LAST = ["Silva", "Nguyen", "Okafor", "Mendes", "Kowalski", "Ferreira", "Haddad", "Lima", "Osei",
        "Barros", "Duarte", "Yilmaz"]
ROUTES = ["Route 12", "Route 44", "Route 7", "Route 31", "Route 19"]

HEADERS = ["Date", "Route", "Trips", "Unit cost", "Total"]


def build(path, staff_id, name, defect=None):
    wb = Workbook()
    ws = wb.active
    ws.title = "Claim"

    ws["A1"] = "MONTHLY TRANSPORT CLAIM"
    ws["A1"].font = Font(name="Arial", size=12, bold=True)
    ws["A2"] = "Staff ID"
    ws["B2"] = staff_id
    ws["A3"] = "Name"
    ws["B3"] = name
    ws["A4"] = "Period"
    ws["B4"] = "2024-06"

    for col, header in enumerate(HEADERS, start=1):
        cell = ws.cell(row=6, column=col, value=header)
        cell.font = Font(name="Arial", size=10, bold=True)

    row = 7
    entries = []
    for day in random.sample(range(1, 29), random.randint(8, 14)):
        entries.append([f"2024-06-{day:02d}", random.choice(ROUTES),
                        random.choice([1, 2]), round(random.uniform(3.5, 9.8), 2)])
    entries.sort()

    if defect == "duplicate":
        entries.append(list(entries[0]))

    for i, (date, route, trips, cost) in enumerate(entries):
        ws.cell(row=row, column=1, value=date)
        ws.cell(row=row, column=2, value=route)

        if defect == "text_number" and i == 2:
            ws.cell(row=row, column=3, value=str(trips))  # typed as text
        else:
            ws.cell(row=row, column=3, value=trips)

        if defect == "missing" and i == 4:
            ws.cell(row=row, column=4, value=None)  # left blank
        else:
            ws.cell(row=row, column=4, value=cost)

        ws.cell(row=row, column=5, value=f"=C{row}*D{row}")
        row += 1

    ws.cell(row=row + 1, column=4, value="Claim total").font = Font(name="Arial", bold=True)
    ws.cell(row=row + 1, column=5, value=f"=SUM(E7:E{row - 1})").font = Font(name="Arial", bold=True)

    for column, width in zip("ABCDE", (14, 14, 8, 12, 12)):
        ws.column_dimensions[column].width = width

    wb.save(path)


def main(count=12, seed=11):
    random.seed(seed)
    INDIVIDUAL.mkdir(exist_ok=True)
    for old in INDIVIDUAL.glob("*.xlsx"):
        old.unlink()

    defects = {2: "missing", 5: "text_number", 9: "duplicate"}

    for i in range(count):
        staff_id = str(random.randint(1000000000, 9999999999))
        name = f"{random.choice(FIRST)} {random.choice(LAST)}"
        path = INDIVIDUAL / f"claim_{staff_id}.xlsx"
        build(path, staff_id, name, defect=defects.get(i))
        flag = f"  [seeded defect: {defects[i]}]" if i in defects else ""
        print(f"created {path.name}{flag}")

    print(f"\n{count} files in {INDIVIDUAL}/ — 3 of them contain deliberate defects")


if __name__ == "__main__":
    main()
